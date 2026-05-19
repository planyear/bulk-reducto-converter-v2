from pathlib import Path

from openpyxl import load_workbook


def convert(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sections: list[str] = []
    try:
        for ws in wb.worksheets:
            sections.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                cells = ["" if v is None else str(v).replace("|", "\\|") for v in row]
                sections.append("| " + " | ".join(cells) + " |")
            sections.append("")
    finally:
        wb.close()
    return "\n".join(sections).rstrip() + "\n"
